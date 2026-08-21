"""Helpers for importing dynamic scheme entries from spreadsheet-like files."""

from __future__ import annotations

import re
from datetime import date, datetime, time
from io import BytesIO
from pathlib import Path

from .models import SchemeEntry


MAX_IMPORT_ROWS = 1000
MAX_IMPORT_FILE_SIZE = 10 * 1024 * 1024
SUPPORTED_EXTENSIONS = {'.xlsx', '.xlsm', '.xls', '.pdf'}


class SchemeImportError(ValueError):
    """An import file cannot be safely converted into scheme rows."""


def _normalise_header(value):
    return re.sub(r'[^a-z0-9]+', '', str(value or '').strip().lower())


def _json_value(value):
    if value is None:
        return ''
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return str(value).strip() if not isinstance(value, (int, float, bool)) else value


STATUS_ALIASES = {
    'announced': SchemeEntry.STATUS_ANNOUNCED,
    'announcedbutnotstarted': SchemeEntry.STATUS_ANNOUNCED,
    'notstarted': SchemeEntry.STATUS_ANNOUNCED,
    'pending': SchemeEntry.STATUS_ANNOUNCED,
    'inprogress': SchemeEntry.STATUS_IN_PROGRESS,
    'ongoing': SchemeEntry.STATUS_IN_PROGRESS,
    'completedtobeinaugurated': SchemeEntry.STATUS_AWAITING_INAUGURATION,
    'completednotinaugurated': SchemeEntry.STATUS_AWAITING_INAUGURATION,
    'readyforinauguration': SchemeEntry.STATUS_AWAITING_INAUGURATION,
    'completedandinaugurated': SchemeEntry.STATUS_INAUGURATED,
    'inaugurated': SchemeEntry.STATUS_INAUGURATED,
}
STATUS_HEADERS = {'status', 'schemestatus', 'progressstatus'}


def _status_value(value, default_status):
    normalised = _normalise_header(value)
    if not normalised:
        return default_status, False
    valid_values = {choice[0] for choice in SchemeEntry.STATUS_CHOICES}
    if value in valid_values:
        return value, False
    mapped = STATUS_ALIASES.get(normalised)
    return (mapped, False) if mapped else (default_status, True)


def _extract_excel_tables(file_bytes, extension):
    if extension == '.xls':
        try:
            import xlrd
        except ImportError as exc:  # pragma: no cover - deployment dependency guard
            raise SchemeImportError('Legacy .xls support is not installed on the server.') from exc

        try:
            workbook = xlrd.open_workbook(file_contents=file_bytes)
            return [
                [[sheet.cell_value(row, col) for col in range(sheet.ncols)] for row in range(sheet.nrows)]
                for sheet in workbook.sheets()
            ]
        except Exception as exc:
            raise SchemeImportError('The Excel file could not be read. Check that it is not damaged or password protected.') from exc

    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - deployment dependency guard
        raise SchemeImportError('Excel support is not installed on the server.') from exc

    try:
        workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        return [[list(row) for row in sheet.iter_rows(values_only=True)] for sheet in workbook.worksheets]
    except Exception as exc:
        raise SchemeImportError('The Excel file could not be read. Check that it is not damaged or password protected.') from exc


def _extract_pdf_tables(file_bytes):
    try:
        import pdfplumber
    except ImportError as exc:  # pragma: no cover - deployment dependency guard
        raise SchemeImportError('PDF support is not installed on the server.') from exc

    try:
        with pdfplumber.open(BytesIO(file_bytes)) as pdf:
            tables = []
            for page in pdf.pages:
                tables.extend(page.extract_tables() or [])
            return tables
    except Exception as exc:
        raise SchemeImportError('The PDF could not be read. Check that it is not damaged or password protected.') from exc


def _header_match_count(row, field_lookup):
    headers = {_normalise_header(cell) for cell in row if _normalise_header(cell)}
    return len(headers.intersection(field_lookup)) + int(bool(headers.intersection(STATUS_HEADERS)))


def _parse_table(table, field_definitions, default_status, row_offset=0):
    field_lookup = {_normalise_header(field): field for field in field_definitions}
    candidates = [
        (index, _header_match_count(row, field_lookup))
        for index, row in enumerate(table[:20])
        if row
    ]
    if not candidates:
        return [], [], [], [], 0

    header_index, match_count = max(candidates, key=lambda candidate: candidate[1])
    if match_count == 0:
        return [], [], [], [], 0

    header_row = table[header_index]
    field_columns = {}
    status_column = None
    unmatched_headers = []
    matched_headers = []

    for column_index, raw_header in enumerate(header_row):
        normalised = _normalise_header(raw_header)
        if not normalised:
            continue
        if normalised in field_lookup and field_lookup[normalised] not in field_columns:
            field = field_lookup[normalised]
            field_columns[field] = column_index
            matched_headers.append(field)
        elif normalised in STATUS_HEADERS and status_column is None:
            status_column = column_index
            matched_headers.append('Status')
        else:
            unmatched_headers.append(str(raw_header).strip())

    parsed_rows = []
    warnings = []
    for table_row_number, row in enumerate(table[header_index + 1:], start=header_index + 2):
        values = {
            field: _json_value(row[column]) if column < len(row) else ''
            for field, column in field_columns.items()
        }
        if not any(str(value).strip() for value in values.values()):
            continue

        raw_status = row[status_column] if status_column is not None and status_column < len(row) else ''
        entry_status, used_fallback = _status_value(raw_status, default_status)
        source_row = row_offset + table_row_number
        if used_fallback:
            warnings.append(
                f'Row {source_row}: unknown status "{raw_status}"; the default status will be used.'
            )
        parsed_rows.append({
            'source_row': source_row,
            'values': {field: values.get(field, '') for field in field_definitions},
            'status': entry_status,
            'status_display': dict(SchemeEntry.STATUS_CHOICES)[entry_status],
        })

    return parsed_rows, matched_headers, unmatched_headers, warnings, len(table)


def parse_scheme_import(uploaded_file, field_definitions, default_status):
    """Return rows and import metadata without writing to the database."""
    filename = uploaded_file.name or ''
    extension = Path(filename).suffix.lower()
    if extension not in SUPPORTED_EXTENSIONS:
        raise SchemeImportError('Choose an Excel (.xlsx, .xlsm, .xls) or PDF (.pdf) file.')
    if uploaded_file.size > MAX_IMPORT_FILE_SIZE:
        raise SchemeImportError('The file is larger than 10 MB.')
    if not field_definitions:
        raise SchemeImportError('This scheme does not have any fields to import into.')

    file_bytes = uploaded_file.read()
    tables = (
        _extract_pdf_tables(file_bytes)
        if extension == '.pdf'
        else _extract_excel_tables(file_bytes, extension)
    )

    all_rows = []
    all_matched_headers = []
    all_unmatched_headers = []
    warnings = []
    row_offset = 0
    for table in tables:
        rows, matched, unmatched, table_warnings, table_length = _parse_table(
            table, field_definitions, default_status, row_offset=row_offset
        )
        all_rows.extend(rows)
        all_matched_headers.extend(matched)
        all_unmatched_headers.extend(unmatched)
        warnings.extend(table_warnings)
        row_offset += table_length
        if len(all_rows) > MAX_IMPORT_ROWS:
            raise SchemeImportError(f'The file contains more than {MAX_IMPORT_ROWS} data rows. Split it into smaller files.')

    if not all_matched_headers:
        expected = ', '.join(field_definitions)
        raise SchemeImportError(f'No matching column headers were found. Expected at least one of: {expected}.')
    if not all_rows:
        raise SchemeImportError('No data rows were found below the matching column headers.')

    return {
        'rows': all_rows,
        'matched_headers': list(dict.fromkeys(all_matched_headers)),
        'unmatched_headers': list(dict.fromkeys(filter(None, all_unmatched_headers))),
        'warnings': warnings,
    }
